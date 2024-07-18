import csv
from datetime import datetime
from openpyxl import load_workbook
from yattag import Doc, indent
from tqdm import tqdm


class Brand:
    def __init__(self, BrandID, BrandName):
        self.BrandID = BrandID
        self.BrandName = BrandName

    def to_xml(self):
        doc, tag, text = Doc().tagtext()
        with tag("Brand"):
            with tag("ExternalId"):
                text(self.BrandID)
            with tag("Name"):
                text(self.BrandName)
        return indent(doc.getvalue())


class Category:
    def __init__(self, CategoryID, CategoryName):
        self.CategoryID = CategoryID
        self.CategoryName = CategoryName

    def to_xml(self):
        doc, tag, text = Doc().tagtext()
        with tag("Category"):
            with tag("ExternalId"):
                text(self.CategoryID)
            with tag("Name"):
                text(self.CategoryName)
        return indent(doc.getvalue())


class Product:
    def __init__(
        self,
        ProductId,
        ProductName,
        ProductPageURL,
        ProductImageURL,
        ProductDescription,
        BrandID,
        CategoryID,
        Inactive,
        ModelNumber,
        ManufacturerPartNumber,
        UPC,
        EAN,
        ProductFamily,
        ProductFamily_Expand,
    ):
        self.ProductID = ProductId
        self.ProductName = ProductName
        self.ProductPageURL = ProductPageURL
        self.ProductImageURL = ProductImageURL
        self.ProductDescription = ProductDescription
        self.Brand = Brand
        self.BrandID = BrandID
        self.Category = Category
        self.CategoryID = CategoryID
        self.Inactive = Inactive
        self.ModelNumber = ModelNumber
        self.ManufacturerPartNumber = ManufacturerPartNumber
        self.UPC = UPC
        self.EAN = EAN
        self.ProductFamily = ProductFamily
        self.ProductFamily_Expand = ProductFamily_Expand

    def to_xml(self):
        doc, tag, text = Doc().tagtext()
        with tag("Product", removed=self.Inactive.lower()):
            with tag("ExternalId"):
                text(self.ProductID)
            with tag("Name"):
                text(self.ProductName)
            with tag("Description"):
                text(self.ProductDescription)
            with tag("BrandExternalId"):
                text(self.BrandID)
            with tag("CategoryExternalId"):
                text(self.CategoryID)
            with tag("ProductPageURL"):
                text(self.ProductPageURL)
            with tag("ImageURL"):
                text(self.ProductImageURL)
            with tag("EANs"):
                for ean in self.EAN.split(","):
                    with tag("EAN"):
                        text(ean)
            with tag("UPCs"):
                for upc in self.UPC.split(","):
                    with tag("UPC"):
                        text(upc)
            with tag("Attributes"):
                with tag("Attribute", id="BV_FE_FAMILY"):
                    with tag("Value"):
                        text(self.ProductFamily)
                with tag("Attribute", id="BV_FE_EXPAND"):
                    with tag("Value"):
                        text(self.ProductFamily_Expand)
        return indent(doc.getvalue())


def read(file):
    if file.endswith(".csv"):
        with open(file, "r", encoding="cp1252") as file:
            data = list(csv.reader(file))
        return data

    elif file.endswith(".xlsx"):
        workbook = load_workbook(file)
        sheet = workbook.active
        data = [list(value) for value in sheet.iter_cols(values_only=True)]
        return data
    else:
        raise ValueError("Invalid file format")


def convert(data):

    print("Converting data to XML format...")
    doc, tag, text = Doc().tagtext()
    xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
    extract_date = str(datetime.today()).split()[0]
    doc.asis(xml_header)
    doc.asis(f"<ExtractDate>{extract_date}</ExtractDate>")

    headers = data[0]
    products = []
    brands = []
    categories = []

    for row in tqdm(data[1:], unit="rows"):

        if row[headers.index("BrandID")] not in [brand for brand in brands]:
            new_brand = Brand(
                row[headers.index("BrandID")], row[headers.index("Brand")]
            )
            brands.append(new_brand.BrandID)
            doc.asis(new_brand.to_xml())

        if row[headers.index("CategoryID")] not in [
            category for category in categories
        ]:
            new_category = Category(
                row[headers.index("CategoryID")], row[headers.index("Category")]
            )
            categories.append(new_category.CategoryID)
            doc.asis(new_category.to_xml())

        if row[headers.index("ProductID")] not in [product for product in products]:
            new_product = Product(
                row[headers.index("ProductID")],
                row[headers.index("ProductName")],
                row[headers.index("ProductPageURL")],
                row[headers.index("ProductImageURL")],
                row[headers.index("ProductDescription")],
                row[headers.index("BrandID")],
                row[headers.index("CategoryID")],
                row[headers.index("Inactive")],
                row[headers.index("ModelNumber")],
                row[headers.index("ManufacturerPartNumber")],
                row[headers.index("UPC")],
                row[headers.index("EAN")],
                row[headers.index("ProductFamily")],
                row[headers.index("ProductFamily-Expand")],
            )

            products.append(new_product.ProductID)
            doc.asis(new_product.to_xml())

    return indent(doc.getvalue())


def write(file, xml_data):
    file = file if file.endswith(".xml") else file + ".xml"

    with open(file, "w", encoding="utf-8") as file:
        file.write(xml_data)

    print(f"File successfully written to {file.buffer.name}")
